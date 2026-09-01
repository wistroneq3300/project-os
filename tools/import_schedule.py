#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ProjectOS — Daily Schedule.xlsx 匯入器
將硬體 Bring-up 排程表轉換成 assets/js/data.js 用的專案資料格式。

Excel 結構(單一工作表 'waterfall'):
  * 時間軸:第 1 列(G 欄起)每天一欄,2026-08-31 ~ 2026-11-27
  * 里程碑:EB1 / TS1 / EB2 / TS2,各有自己的日期列 + 任務區塊
  * A 欄 = 群組(PCBA 區:Baseboard/... ; Function 區:Mechenical/...)
  * B 欄 = 任務名稱, D/E = 開始/結束日期, F = 註記(週數)
  * 任務列上對應日期欄的 '1' = 那天有排程

輸出:每個里程碑一個「專案」,任務保留原名稱與群組,推估日期/工作量。
"""
import json
import re
from datetime import date, datetime, timedelta
from openpyxl import load_workbook

SRC = "Daily Schedule.xlsx"
XLSX_PATH = SRC
OUT = "assets/js/data.js"

def read_axis(ws):
    """回傳 {column_index: date} 從第 1 列讀出時間軸。"""
    axis = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, datetime):
            axis[c] = v.date()
    return axis

def col_for_date(axis, d):
    for c, dt in axis.items():
        if dt == d:
            return c
    return None

def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["waterfall"]
    axis = read_axis(ws)              # {col: date}
    dates = sorted(axis.values())     # sorted date list
    d0, d1 = dates[0], dates[-1]
    n_days = (d1 - d0).days + 1

    # ---------------- 里程碑/區塊解析 ----------------
    # 用 A 欄的值搭配日期列偵測里程碑;每個里程碑有自己的 header 列
    milestones = []   # list of {name, blocks:[...]}
    current_milestone = None

    # 依列掃描,先建立「里程碑 → 任務列範圍」
    rows = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        # 該列是否有任務名稱
        is_task = bool(b) and str(b).strip() not in ('Task',)
        # 該列是否有日期欄標記 (1)
        marks = [axis[c] for c in axis if ws.cell(row=r, column=c).value not in (None, '')]
        rows.append(dict(r=r, a=a, b=b, d=d, e=e, f=f,
                         is_task=is_task, marks=sorted(set(marks))))

    GROUP_HEADERS = {'PCBA', 'Function', 'Task', 'EB1', 'EB2', 'TS1', 'TS2'}
    MILESTONES = ('EB1', 'EB2', 'TS1', 'TS2')

    # 一次掃描:當 A 欄出現里程碑名 → 切換目前里程碑;出現群組名 → 更新目前群組;
    # 列有任務名 → 記入目前里程碑(task_rows 帶有此刻的群組)。
    current_group = '未分組'
    for row in rows:
        a = row['a']
        if row['r'] == 1:
            current_milestone = {'name': str(a).strip(), 'task_rows': []}
            milestones.append(current_milestone)
            continue
        if isinstance(a, str):
            a_clean = a.strip()
            if a_clean in MILESTONES:
                current_milestone = {'name': a_clean, 'task_rows': []}
                milestones.append(current_milestone)
                continue
            if a_clean not in GROUP_HEADERS and a_clean != '未分組':
                current_group = a_clean
        if current_milestone is not None and row['is_task']:
            row['group'] = current_group
            current_milestone['task_rows'].append(row)

    # ---------------- 日期排程 ----------------
    # 週數註記 -> 天數(作為該任務的「相對工作量」)
    def weeks_from_note(note):
        if not note:
            return None
        m = re.search(r'([\d.]+)\s*周', str(note))
        if m:
            return float(m.group(1)) * 7.0
        return None

    # 對缺少完整日期的任務,在里程碑時間窗 [ms_base, d1] 內依序平攤。
    # 有 D 的用真實 D(不超過 d1);其餘依週數/預設時長排入。
    # 所有 end 都 clamp 到 d1,確保全部落在時間軸內、甘特圖可讀。

    DEFAULT_DUR = 5
    MILESTONE_BASE = {'EB1': 0, 'TS1': 6, 'TS2': 40}   # 里程碑起始(天)偏移

    built_milestones = []
    for ms in milestones:
        name = ms['name']
        ms_d0 = d0 + timedelta(days=MILESTONE_BASE.get(name, 0))

        # 同群組的任務視為「可並行」工作(各自獨立一行、可重疊時間),
        # 因此群組內任務從「群組起始日」各自開始,不互相順序推擠。
        # 不同群組則從里程碑基準錯開起始,讓長測試也能排進時間軸。
        group_start = {}          # group -> 該群組可開始日
        group_idx = {}            # 群組出現順序
        tasks = []
        for row in ms['task_rows']:
            g = row['group']
            label = str(row['b']).strip()
            note = row['f'] or ''

            if g not in group_start:
                group_start[g] = ms_d0 + timedelta(days=len(group_start) * 8)  # 群組間錯開
                group_idx[g] = len(group_start) - 1

            # ---- 起始日:有 D / marks 用真實,否則用群組起始 ----
            if isinstance(row['d'], datetime):
                start = row['d'].date()
            elif row['marks']:
                start = row['marks'][0]
            else:
                start = group_start[g]

            # ---- 持續日數 ----
            if isinstance(row['e'], datetime):
                end = row['e'].date()
            elif row['marks'] and len(row['marks']) > 1:
                end = row['marks'][-1]
            elif row['marks'] and len(row['marks']) == 1:
                end = start                                    # 單日
            else:
                dur = weeks_from_note(note)
                if dur is None:
                    dur = DEFAULT_DUR
                end = start + timedelta(days=int(dur) - 1)

            # ---- clamp 到整體時間窗 ----
            if start < d0:   start = d0
            if end < start:  end = start
            if start > d1:   start = d1 - timedelta(days=max(DEFAULT_DUR, int(dur if weeks_from_note(note) else 0)))
            if end > d1:     end = d1

            # ---- 狀態 ----
            status, progress = 'todo', 0
            if 'Early' in label:
                status, progress = 'done', 100
            elif 'Bring up in SC' in label or 'full qual' in label or 'Cert' in label:
                status, progress = 'doing', 30

            tasks.append(dict(name=label, group=g, start=start, end=end,
                              status=status, progress=progress, note=note))

        built_milestones.append(dict(name=name, tasks=tasks))

    # 儲存 debug 中間檔(供檢查)
    debug = {
        'window': [d0.isoformat(), d1.isoformat()],
        'n_days': n_days,
        'milestones': [
            {'name': m['name'],
             'tasks': [{'name': t['name'], 'group': t['group'],
                        'start': t['start'].isoformat(), 'end': t['end'].isoformat(),
                        'status': t['status'], 'progress': t['progress']}
                       for t in m['tasks']]}
            for m in built_milestones
        ],
    }
    with open('tools/schedule_debug.json', 'w', encoding='utf-8') as fh:
        json.dump(debug, fh, ensure_ascii=False, indent=2)

    # ---------------- 產生 assets/js/data.js ----------------
    import json as _json

    MILESTONE_META = {
        # name: (中文名, color, budget_base_each)
        'EB1': ('EB1 工程驗證', '#ffb224', 180000),
        'TS1': ('TS1 機種測試', '#60a5fa', 260000),
        'EB2': ('EB2 工程驗證', '#a78bfa', 180000),
        'TS2': ('TS2 機種測試', '#34d399', 260000),
    }
    OWNER_KEYS = ['alan', 'may', 'john', 'soph', 'tom']
    OWNER_LOOP = 0

    def iso(d):
        return f"{d.year}-{str(d.month).zfill(2)}-{str(d.day).zfill(2)}"

    # 過濾掉空里程碑
    active = [m for m in built_milestones if m['tasks']]

    # 每群組計次,產生唯一 task id
    group_counter = {}
    projects_js = []
    for ms in active:
        nm = ms['name']
        disp, color, base_per_task = MILESTONE_META[nm]
        tasks = []
        deps = []
        n = len(ms['tasks'])
        for i, t in enumerate(ms['tasks']):
            g = t['group']
            t_id = f"{nm.lower()}-{i+1}"
            owner = OWNER_KEYS[OWNER_LOOP % len(OWNER_KEYS)]
            OWNER_LOOP += 1
            tasks.append(dict(
                id=t_id,
                name=f"{t['name']}" if g == '未分組' else f"{g} · {t['name']}",
                grp=g,
                status=t['status'],
                owner=owner,
                start=iso(t['start']),
                end=iso(t['end']),
                progress=t['progress'],
            ))
            # 群組內 deps:同一群組前一任務 → 下一任務
            if tasks:                                # tasks 已含先前處理的任務
                prev = tasks[-1]
                if prev['grp'] == g:
                    deps.append([prev['id'], t_id])

        spend_ratio = 0.42   # 已投入比例(示範)
        budget = n * base_per_task
        spent = int(budget * spend_ratio)
        proj_status = 'done' if all(t['status'] == 'done' for t in tasks) else (
            'doing' if any(t['status'] == 'doing' for t in tasks) else 'todo')

        projects_js.append(dict(
            id=nm.lower(),
            name=disp,
            color=color,
            status=proj_status,
            budget=budget,
            spent=spent,
            manager=OWNER_KEYS[OWNER_LOOP % len(OWNER_KEYS)],
            source_milestone=nm,
            tasks=tasks,
            deps=deps,
        ))
        OWNER_LOOP += 1

    # 序列化成 data.js
    header = (
        "/* ============================================================\n"
        "   ProjectOS — 資料來源:Daily Schedule.xlsx(硬體 Bring-up 排程)\n"
        "   由 tools/import_schedule.py 自動產生,請勿手動編輯。\n"
        "   手動改資料請重跑:python3 tools/import_schedule.py\n"
        "   ============================================================ */\n"
    )
    body = _json.dumps(projects_js, ensure_ascii=False, indent=2)

    status_js = (
        "const STATUS_META = {\n"
        "  todo:   { label: '尚未開始', color: '#5d6b7e', bg: 'rgba(93,107,126,.18)' },\n"
        "  doing:  { label: '進行中',   color: '#60a5fa', bg: 'rgba(96,165,250,.15)' },\n"
        "  block:  { label: '卡關',     color: '#f87171', bg: 'rgba(248,113,113,.14)' },\n"
        "  done:   { label: '已完成',   color: '#34d399', bg: 'rgba(52,211,153,.14)' },\n"
        "};\n\n"
        "const OWNERS = {\n"
        "  alan: { name: 'Alan',  color: '#60a5fa' },\n"
        "  may:  { name: 'May',   color: '#f472b6' },\n"
        "  john: { name: 'John',  color: '#34d399' },\n"
        "  soph: { name: 'Soph',  color: '#fbbf24' },\n"
        "  tom:  { name: 'Tom',   color: '#a78bfa' },\n"
        "};\n\n"
        "const PROJECTS = \n"
    )
    helper_js = (
        "\n"
        "const projById = id => PROJECTS.find(p => p.id === id);\n"
        "const taskById = id => { for (const p of PROJECTS) { const t = p.tasks.find(t => t.id === id); if (t) return t; } return null; };\n"
        "function projProgress(p){ if (!p.tasks.length) return 0; return Math.round(p.tasks.reduce((s,t)=>s+t.progress,0)/p.tasks.length); }\n"
        "const daySpan = (a,b) => Math.round((new Date(b)-new Date(a))/86400000);\n"
    )
    data_js = header + "\n" + status_js + body + ";\n" + helper_js
    with open(OUT, 'w', encoding='utf-8') as fh:
        fh.write(data_js)
    print(f"\n✅ 已寫出 {OUT} ({len(projects_js)} 個里程碑專案)")
    print(f"   總任務數: {sum(len(p['tasks']) for p in projects_js)}")

if __name__ == '__main__':
    main()
